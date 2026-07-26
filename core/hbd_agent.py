#!/usr/bin/env python3
"""
HBD
Hooper-Borg-Dergaa Agent (HBD): deterministic core behind the HBD application.
Daily training-load and wellness monitoring with next-day decisions, for any sport.

Pipeline:
  1. Ingest daily self-report exported from the Google Form responses sheet.
  2. Compute session-RPE load, daily/weekly load, monotony, strain, ACWR,
     and a direction-harmonised wellness composite, per athlete.
  3. Decide next-day training: INCREASE, DECREASE or MAINTAIN, with a rationale.
  4. Produce a coach email (HTML + plain text) and append an audit log so every
     recommendation is timestamped and reproducible.

Design rule: all numbers are computed deterministically here. A language model,
if used downstream, only rephrases this output. It never changes a metric.

Author:  Dr. Ismail Dergaa et al. License: MIT.
"""

import argparse
import csv
import io
import math
import os
import re
import statistics as st
from datetime import datetime, date, timedelta

# ----------------------------------------------------------------------------
# Configuration. Thresholds are explicit so the decision logic is auditable.
# ----------------------------------------------------------------------------
ACUTE_DAYS = 7
CHRONIC_DAYS = 28
WELLNESS_BASELINE_DAYS = 28

# Acute:chronic workload ratio reference band. Treated as a prompt for judgement,
# not a verdict (see Hulin & Gabbett, BJSM 2019).
ACWR_LOW = 0.80      # below this: room to add load
ACWR_HIGH = 1.30     # above this: caution
ACWR_RED = 1.50      # above this: clear overload signal

MONOTONY_HIGH = 2.0  # Foster 1998: high monotony with high load raises risk
WELLNESS_DROP_Z = -1.0   # today's wellness this far below personal baseline = red
WELLNESS_GOOD_Z = 0.0    # at or above baseline = green
ACUTE_SPIKE_RATIO = 1.5  # today's load vs 7-day mean

# ----------------------------------------------------------------------------
# Column mapping. Matches French and English form headers (substring, lower-case).
# ----------------------------------------------------------------------------
COLMAP = {
    "timestamp": ["horodat", "timestamp"],
    "email": ["email", "adresse", "utilisateur", "mail", "athlete id"],
    "athlete": ["qui es", "player", "nageur", "swimmer", "athlete", "name", "nom"],
    "moment": ["moment", "timing"],
    "duration": ["duree", "durée", "duration", "min"],
    "intensity": ["intensite", "intensité", "intensity", "cr-10", "ressenti", "rpe"],
    "fatigue": ["fatigu"],
    "stress": ["stress"],
    "soreness": ["courbatur", "muscle", "soreness"],
    "sleep": ["sommeil", "sleep"],
    # Health / injury module (HBD daily pain gate + weekly OSTRC-H)
    "pain": ["douleur", "gêne", "gene specifique", "pain", "health problem"],
    "pain_location": ["localis", "location", "zone"],
    "ostrc_q1": ["ostrc_q1", "participation"],
    "ostrc_q2": ["ostrc_q2", "volume"],
    "ostrc_q3": ["ostrc_q3", "performance ostrc", "ostrc performance"],
    "ostrc_q4": ["ostrc_q4", "symptom"],
}


def _match_col(header_row):
    """Return {canonical: actual_header}. Key-priority match, one header per field.
    Email is matched before athlete so 'Nom d'utilisateur' is not taken as the name."""
    mapping = {}
    used = set()
    lowered = [(h, h.lower().strip()) for h in header_row]
    for canon, keys in COLMAP.items():
        for k in keys:                       # try keys in priority order
            for h, hl in lowered:
                if h in used:
                    continue
                if k in hl:
                    mapping[canon] = h
                    used.add(h)
                    break
            if canon in mapping:
                break
    return mapping


def _to_float(x):
    if x is None:
        return None
    s = str(x).strip().replace(",", ".")
    if s == "":
        return None
    # handle values like "5 - Beaucoup" by taking the leading number
    num = ""
    for ch in s:
        if ch.isdigit() or ch == "." or (ch == "-" and num == ""):
            num += ch
        elif num:
            break
    try:
        return float(num)
    except ValueError:
        return None


def _yesno(x):
    """Map a free-text yes/no (French, Arabic, English) to True/False/None."""
    if x is None:
        return None
    s = str(x).strip().lower()
    if s == "":
        return None
    if s[0] in ("o", "y", "1") or "oui" in s or "yes" in s or "نعم" in s:
        return True
    if s[0] in ("n", "0") or "non" in s or "no" in s or "لا" in s:
        return False
    return None


def _parse_date(x):
    if not x:
        return None
    s = str(x).strip()
    s = re.sub(r"\s+(UTC|GMT)[+-]?\d*", "", s).strip()   # drop trailing timezone, e.g. "UTC+1"
    for fmt in ("%Y/%m/%d %I:%M:%S %p", "%m/%d/%Y %I:%M:%S %p", "%d/%m/%Y %I:%M:%S %p",
                "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
                "%m/%d/%Y %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y",
                "%Y-%m-%d %I:%M:%S %p"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # last resort: take first 10 chars as ISO
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


# ----------------------------------------------------------------------------
# Ingestion
# ----------------------------------------------------------------------------
def _read_rows(path):
    """Read a CSV or an Excel (.xlsx) export of the form responses into a list of rows."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError("Reading Excel needs openpyxl: pip install openpyxl") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in r])
        wb.close()
        return rows
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def load_records(path):
    rows = _read_rows(path)
    if not rows:
        return []
    header = rows[0]
    cmap = _match_col(header)
    required = ["athlete", "duration", "intensity"]
    missing = [r for r in required if r not in cmap]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Found: {list(cmap)}")
    idx = {c: header.index(h) for c, h in cmap.items()}
    recs = []
    for r in rows[1:]:
        if not any(cell.strip() for cell in r):
            continue
        def g(c):
            return r[idx[c]] if c in idx and idx[c] < len(r) else None
        d = _parse_date(g("timestamp"))
        rec = {
            "date": d,
            "athlete": (g("athlete") or "").strip(),
            "moment": (g("moment") or "").strip(),
            "duration": _to_float(g("duration")),
            "intensity": _to_float(g("intensity")),
            "fatigue": _to_float(g("fatigue")),
            "stress": _to_float(g("stress")),
            "soreness": _to_float(g("soreness")),
            "sleep": _to_float(g("sleep")),
            "pain": _yesno(g("pain")),
            "pain_location": (g("pain_location") or "").strip() or None,
            "ostrc_q1": _to_float(g("ostrc_q1")),
            "ostrc_q2": _to_float(g("ostrc_q2")),
            "ostrc_q3": _to_float(g("ostrc_q3")),
            "ostrc_q4": _to_float(g("ostrc_q4")),
        }
        if rec["athlete"] and rec["duration"] is not None and rec["intensity"] is not None:
            rec["session_load"] = rec["duration"] * rec["intensity"]
        else:
            rec["session_load"] = None
        recs.append(rec)
    return recs


# ----------------------------------------------------------------------------
# Per-athlete daily aggregation
# ----------------------------------------------------------------------------
def wellness_composite(rec):
    """Harmonise direction so higher = better. Range ~1..7. None if no items.
    All four Hooper items are recorded 1=best..7=worst (sleep 1=very very good),
    so every item is flipped (8 - x) to make the composite higher = better."""
    vals = []
    for k in ("sleep", "fatigue", "stress", "soreness"):
        if rec[k] is not None:
            vals.append(8 - rec[k])            # 1=best..7=worst -> flip so higher = better
    return round(sum(vals) / len(vals), 2) if vals else None


def daily_series(recs, athlete):
    """Return {date: {'load':, 'wellness':}} summing load, averaging wellness."""
    by_day = {}
    for r in recs:
        if r["athlete"] != athlete or r["date"] is None:
            continue
        d = by_day.setdefault(r["date"], {"loads": [], "well": []})
        if r["session_load"] is not None:
            d["loads"].append(r["session_load"])
        w = wellness_composite(r)
        if w is not None:
            d["well"].append(w)
    out = {}
    for d, v in by_day.items():
        out[d] = {
            "load": sum(v["loads"]) if v["loads"] else 0.0,
            "wellness": round(sum(v["well"]) / len(v["well"]), 2) if v["well"] else None,
        }
    return out


def window_loads(series, end, days):
    """Daily loads across every calendar day in the window. Missing/rest days = 0,
    as required for Foster monotony and strain and for ACWR baselines."""
    start = end - timedelta(days=days - 1)
    out = []
    d = start
    while d <= end:
        out.append(series[d]["load"] if d in series else 0.0)
        d += timedelta(days=1)
    return out


def compute_metrics(series, ref_date):
    """Compute the load/wellness indices for an athlete on ref_date."""
    if ref_date not in series:
        return None
    acute_window = window_loads(series, ref_date, ACUTE_DAYS)
    # Uncoupled ACWR: chronic window excludes the current acute period to avoid
    # mathematical coupling between numerator and denominator.
    chronic_window = window_loads(series, ref_date - timedelta(days=ACUTE_DAYS), CHRONIC_DAYS)

    daily_load = series[ref_date]["load"]
    weekly_load = sum(acute_window)
    acute_mean = (sum(acute_window) / len(acute_window)) if acute_window else 0.0
    chronic_daily_mean = (sum(chronic_window) / len(chronic_window)) if chronic_window else 0.0

    # Monotony / strain over the acute (7-day) window
    if len(acute_window) >= 2 and st.pstdev(acute_window) > 0:
        monotony = acute_mean / st.pstdev(acute_window)
    else:
        monotony = None
    strain = weekly_load * monotony if monotony is not None else None

    # ACWR (uncoupled): acute mean daily / chronic mean daily (acute excluded)
    acwr = (acute_mean / chronic_daily_mean) if chronic_daily_mean > 0 else None

    # Acute spike: today vs 7-day mean (excluding today if possible)
    prior = [l for d, l in [(d, series[d]["load"]) for d in series
             if ref_date - timedelta(days=ACUTE_DAYS) <= d < ref_date]]
    prior_mean = (sum(prior) / len(prior)) if prior else 0.0
    acute_spike = (daily_load / prior_mean) if prior_mean > 0 else None

    # Wellness baseline z-score
    wb = [series[d]["wellness"] for d in series
          if ref_date - timedelta(days=WELLNESS_BASELINE_DAYS) <= d < ref_date
          and series[d]["wellness"] is not None]
    today_well = series[ref_date]["wellness"]
    if today_well is not None and len(wb) >= 3 and st.pstdev(wb) > 0:
        well_z = (today_well - (sum(wb) / len(wb))) / st.pstdev(wb)
    else:
        well_z = None

    # Distinct days the athlete actually logged in the last 28 days, for a data-confidence tier.
    days_logged = sum(1 for d in series
                      if ref_date - timedelta(days=CHRONIC_DAYS) < d <= ref_date)
    if days_logged >= 14:
        confidence = "high"
    elif days_logged >= 7:
        confidence = "moderate"
    else:
        confidence = "low"

    return {
        "daily_load": round(daily_load, 1),
        "weekly_load": round(weekly_load, 1),
        "monotony": round(monotony, 2) if monotony is not None else None,
        "strain": round(strain, 1) if strain is not None else None,
        "acwr": round(acwr, 2) if acwr is not None else None,
        "acute_spike": round(acute_spike, 2) if acute_spike is not None else None,
        "wellness": today_well,
        "wellness_z": round(well_z, 2) if well_z is not None else None,
        "days_of_data": len(chronic_window),
        "days_logged": days_logged,
        "confidence": confidence,
    }


# ----------------------------------------------------------------------------
# Decision logic
# ----------------------------------------------------------------------------
def decide(m):
    """Return (decision, rationale_list). DECREASE / MAINTAIN / INCREASE."""
    reasons = []
    red = []
    green = []

    if m["acwr"] is not None:
        if m["acwr"] >= ACWR_RED:
            red.append(f"ACWR {m['acwr']} is high (>= {ACWR_RED}), workload spike")
        elif m["acwr"] > ACWR_HIGH:
            red.append(f"ACWR {m['acwr']} above the {ACWR_HIGH} caution band")
        elif m["acwr"] < ACWR_LOW:
            green.append(f"ACWR {m['acwr']} below {ACWR_LOW}, room to load")

    if m["wellness_z"] is not None:
        if m["wellness_z"] <= WELLNESS_DROP_Z:
            red.append(f"wellness {m['wellness_z']} SD below personal baseline")
        elif m["wellness_z"] >= WELLNESS_GOOD_Z:
            green.append("wellness at or above personal baseline")

    if m["monotony"] is not None and m["monotony"] >= MONOTONY_HIGH:
        red.append(f"high monotony ({m['monotony']}), low day-to-day variation")

    if m["acute_spike"] is not None and m["acute_spike"] >= ACUTE_SPIKE_RATIO:
        red.append(f"acute load spike ({m['acute_spike']}x recent mean)")

    if m.get("confidence") == "low":
        reasons.append(f"limited history ({m.get('days_logged', '?')} days logged); interpret with caution")

    if red:
        decision = "DECREASE"
        reasons = red + reasons
    elif green and len(green) >= 2:
        decision = "INCREASE"
        reasons = green + reasons
    else:
        decision = "MAINTAIN"
        reasons = (green or ["metrics within normal range"]) + reasons
    return decision, reasons


def strategy(decision, m):
    """Map a decision and its dominant signal to one concrete coaching strategy.
    Deterministic: same inputs always give the same suggestion. Advisory only."""
    low_wellness = m.get("wellness_z") is not None and m["wellness_z"] <= WELLNESS_DROP_Z
    high_acwr = m.get("acwr") is not None and m["acwr"] > ACWR_HIGH
    spike = m.get("acute_spike") is not None and m["acute_spike"] >= ACUTE_SPIKE_RATIO
    monotonous = m.get("monotony") is not None and m["monotony"] >= MONOTONY_HIGH
    if decision == "DECREASE":
        if low_wellness:
            return ("Recovery emphasis: cut total volume by about 30 percent, keep intensity light, "
                    "protect sleep and add active recovery, then recheck wellness before the next session.")
        if spike or high_acwr:
            return ("Deload the spike: reduce session volume by 20 to 30 percent and cap intensity, "
                    "hold for one to two days until the workload ratio settles back toward the optimal band.")
        if monotonous:
            return ("Break the monotony: vary the load across the week, insert one clearly easy and one harder day "
                    "instead of repeating a uniform load.")
        return "Ease the next session: lower volume or intensity and reassess tomorrow."
    if decision == "INCREASE":
        return ("Progress within limits: add load gradually, keep the week-to-week ramp under about 10 percent, "
                "and re-evaluate wellness before the next step up.")
    return "Hold the plan: keep the prescribed load and keep monitoring; no change indicated today."


# ----------------------------------------------------------------------------
# Reporting
# ----------------------------------------------------------------------------
BADGE = {"INCREASE": "#1a7f37", "MAINTAIN": "#9a6700", "DECREASE": "#b42318"}


def build_report(results, ref_date):
    n = len(results)
    counts = {"INCREASE": 0, "MAINTAIN": 0, "DECREASE": 0}
    for r in results:
        counts[r["decision"]] += 1
    tomorrow = ref_date + timedelta(days=1)

    rows_html = ""
    for r in sorted(results, key=lambda x: ["DECREASE", "MAINTAIN", "INCREASE"].index(x["decision"])):
        m = r["metrics"]
        rows_html += f"""
        <tr>
          <td><b>{r['athlete']}</b></td>
          <td><span style="background:{BADGE[r['decision']]};color:#fff;padding:2px 8px;border-radius:10px;font-size:12px">{r['decision']}</span></td>
          <td>{m['acwr'] if m['acwr'] is not None else '-'}</td>
          <td>{m['wellness'] if m['wellness'] is not None else '-'} (z {m['wellness_z'] if m['wellness_z'] is not None else '-'})</td>
          <td>{m['monotony'] if m['monotony'] is not None else '-'}</td>
          <td>{m['weekly_load']}</td>
          <td style="font-size:12px;color:#444">{'; '.join(r['reasons'])}</td>
        </tr>"""

    html = f"""<!DOCTYPE html><html><body style="font-family:Arial,Helvetica,sans-serif;color:#1a1a1a;max-width:880px">
    <h2 style="color:#1F3864;margin-bottom:2px">HBD daily report</h2>
    <p style="color:#555;margin-top:0">Data through {ref_date.isoformat()} &middot; recommendation for {tomorrow.isoformat()}</p>
    <p>Athletes analysed: <b>{n}</b> &nbsp;|&nbsp;
       <span style="color:{BADGE['DECREASE']}">Decrease: {counts['DECREASE']}</span> &nbsp;
       <span style="color:{BADGE['MAINTAIN']}">Maintain: {counts['MAINTAIN']}</span> &nbsp;
       <span style="color:{BADGE['INCREASE']}">Increase: {counts['INCREASE']}</span></p>
    <table cellspacing="0" cellpadding="6" style="border-collapse:collapse;width:100%;font-size:13px">
      <tr style="background:#DDEBF7;text-align:left">
        <th>Athlete</th><th>Tomorrow</th><th>ACWR</th><th>Wellness</th><th>Monotony</th><th>Weekly load (AU)</th><th>Why</th>
      </tr>{rows_html}
    </table>
    <p style="font-size:12px;color:#777;margin-top:14px">
      Recommendations are decision support, not verdicts. ACWR and related indices flag patterns for the coach to judge.
      Increase = room to progress load. Decrease = ease tomorrow. Maintain = hold the plan.
      Generated automatically by HBD.</p>
    </body></html>"""

    # plain text
    lines = [f"HBD daily report - data through {ref_date.isoformat()}",
             f"Recommendation for {tomorrow.isoformat()}",
             f"Decrease {counts['DECREASE']} | Maintain {counts['MAINTAIN']} | Increase {counts['INCREASE']}", ""]
    for r in sorted(results, key=lambda x: ["DECREASE", "MAINTAIN", "INCREASE"].index(x["decision"])):
        m = r["metrics"]
        lines.append(f"{r['athlete']}: {r['decision']} | ACWR {m['acwr']} | wellness {m['wellness']} "
                     f"(z {m['wellness_z']}) | monotony {m['monotony']} | weekly {m['weekly_load']} AU")
        lines.append(f"    reason: {'; '.join(r['reasons'])}")
    text = "\n".join(lines)
    return html, text


def append_audit(log_path, results, ref_date):
    new = not os.path.exists(log_path)
    with open(log_path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if new:
            w.writerow(["generated_at", "data_through", "recommend_for", "athlete",
                        "decision", "acwr", "wellness", "wellness_z", "monotony",
                        "strain", "weekly_load", "daily_load", "reasons"])
        gen = datetime.now().isoformat(timespec="seconds")
        rec_for = (ref_date + timedelta(days=1)).isoformat()
        for r in results:
            m = r["metrics"]
            w.writerow([gen, ref_date.isoformat(), rec_for, r["athlete"], r["decision"],
                        m["acwr"], m["wellness"], m["wellness_z"], m["monotony"],
                        m["strain"], m["weekly_load"], m["daily_load"],
                        " | ".join(r["reasons"])])


# ----------------------------------------------------------------------------
# Wellness analytics: sleep, soreness, fatigue (Hooper), alongside load
# ----------------------------------------------------------------------------
def component_daily(recs, athlete):
    """Per-day mean of each Hooper item for one athlete."""
    by = {}
    for r in recs:
        if r["athlete"] != athlete or r["date"] is None:
            continue
        d = by.setdefault(r["date"], {"sleep": [], "fatigue": [], "stress": [], "soreness": []})
        for k in ("sleep", "fatigue", "stress", "soreness"):
            if r[k] is not None:
                d[k].append(r[k])
    return {d: {k: (round(sum(v[k]) / len(v[k]), 1) if v[k] else None) for k in v} for d, v in by.items()}


def _recent(comp, end, days, key):
    start = end - timedelta(days=days - 1)
    return [c[key] for day, c in comp.items() if start <= day <= end and c.get(key) is not None]


def wellness_flags(comp, ref):
    """Plain-language flags from the Hooper items, so wellness, not only load, drives the call."""
    flags = []
    today = comp.get(ref, {})
    sl = _recent(comp, ref, 7, "sleep")   # sleep 1=very good .. 7=very poor
    if today.get("sleep") is not None and today["sleep"] >= 5:
        flags.append(f"poor sleep last night ({today['sleep']}/7)")
    elif sl and sum(sl) / len(sl) >= 4.5:
        flags.append("poor sleep across the week")
    if today.get("fatigue") is not None and today["fatigue"] >= 6:
        flags.append(f"high fatigue ({today['fatigue']}/7)")
    if today.get("stress") is not None and today["stress"] >= 6:
        flags.append(f"high stress ({today['stress']}/7)")
    return flags


# ----------------------------------------------------------------------------
# Targeted help for the coach: each wellness flag maps to a concrete, evidence-based
# behavioural suggestion. Advisory only; anything clinical goes to the medical staff.
# ----------------------------------------------------------------------------
SLEEP_TIP = ("poor sleep: prioritise sleep hygiene, limit bright and blue-light screens for about two hours "
             "before bed, keep a consistent bedtime and a dark, cool, quiet room, and protect the sleep window "
             "around travel and evening matches; any sleep aid or melatonin only on qualified medical advice")
STRESS_TIP = ("high stress: check school, home, and competition pressures, lighten non-training demands where "
              "possible, and consider psychological support")
FATIGUE_TIP = ("high fatigue: favour a recovery day, extend the sleep opportunity, and reassess wellness before "
               "loading again")


def wellness_advice(reasons):
    """Turn wellness flags into concrete coach guidance. Returns a list of tips."""
    tips = []
    joined = " ".join(reasons).lower()
    if "poor sleep" in joined:
        tips.append(SLEEP_TIP)
    if "high stress" in joined:
        tips.append(STRESS_TIP)
    if "high fatigue" in joined:
        tips.append(FATIGUE_TIP)
    return tips


def soreness_reason(comp, series, ref):
    """Explain muscle soreness: a recent load spike points to DOMS; persistence points to a niggle."""
    today = comp.get(ref, {})
    if today.get("soreness") is None or today["soreness"] < 5:
        return None
    prior = [series[d]["load"] for d in series if ref - timedelta(days=7) <= d < ref]
    pm = (sum(prior) / len(prior)) if prior else 0.0
    tl = series.get(ref, {}).get("load", 0.0)
    if pm > 0 and tl >= 1.3 * pm:
        return f"soreness likely from a recent load spike (DOMS), ease tomorrow ({today['soreness']}/7)"
    days_high = sum(1 for d in series if ref - timedelta(days=3) <= d <= ref and comp.get(d, {}).get("soreness", 0) >= 5)
    if days_high >= 3:
        return f"soreness persistent over several days, check for a niggle ({today['soreness']}/7)"
    return f"soreness elevated, monitor ({today['soreness']}/7)"


# ----------------------------------------------------------------------------
# Health / injury module: daily pain gate + weekly OSTRC-H severity
#   OSTRC severity scored 0..100 following the OSTRC method (Clarsen 2014;
#   per-option scoring 0/8/17/25 to be confirmed against Clarsen 2013 before print).
# ----------------------------------------------------------------------------
def health_status(recs, athlete, ref_date):
    """Return the athlete's health signal on ref_date:
    the daily pain flag and location, and the most recent weekly OSTRC-H severity.
    A substantial problem follows Clarsen: a moderate or worse reduction in
    participation, training volume, or performance (any OSTRC item >= 17)."""
    mine = [r for r in recs if r["athlete"] == athlete and r["date"] is not None]
    today = [r for r in mine if r["date"] == ref_date]
    pain = any(r.get("pain") is True for r in today)
    location = next((r.get("pain_location") for r in today
                     if r.get("pain") is True and r.get("pain_location")), None)
    # most recent OSTRC within the last 7 days (weekly instrument)
    ost = [r for r in mine
           if ref_date - timedelta(days=7) <= r["date"] <= ref_date
           and any(r.get(f"ostrc_q{i}") is not None for i in (1, 2, 3, 4))]
    ost.sort(key=lambda r: r["date"])
    severity = None; substantial = False; cannot = False
    if ost:
        r = ost[-1]
        qs = [r.get(f"ostrc_q{i}") or 0 for i in (1, 2, 3, 4)]
        severity = round(sum(qs), 0)
        substantial = any(q >= 17 for q in qs[:3])   # participation, volume, performance
        cannot = any(q >= 25 for q in qs[:3])
    return {"pain": pain, "location": location, "severity": severity,
            "substantial": substantial, "cannot": cannot,
            "flagged": bool(pain or substantial or (severity or 0) > 0)}


def apply_health_override(decision, strat, reasons, hs):
    """Health sits above load and wellness, in three graded tiers. The tool never
    advises loading an athlete who flagged a health problem."""
    loc = f" ({hs['location']})" if hs["location"] else ""
    sev = f", OSTRC severity {int(hs['severity'])}" if hs["severity"] is not None else ""
    hs["action"] = None
    if hs["cannot"]:                                     # OSTRC 'could not participate'
        hs["action"] = "STOP"
        reasons = [f"could not participate{loc}{sev}: STOP, do not train, urgent medical assessment"] + reasons
        return "DECREASE", ("Immediate stop: no training today; arrange an urgent medical assessment "
                            "before any return to load."), reasons
    if hs["substantial"]:                                # moderate or worse reduction
        hs["action"] = "REFER"
        reasons = [f"substantial health problem{loc}{sev}: reduce load and refer to medical staff"] + reasons
        return "DECREASE", ("Reduce and refer: cut load and refer the athlete to the medical staff; "
                            "do not progress until the problem resolves."), reasons
    if hs["flagged"]:                                    # a flag or a minor niggle
        hs["action"] = "HOLD"
        reasons = [f"health flag today{loc}: hold, do not increase, monitor"] + reasons
        if decision == "INCREASE":
            return "MAINTAIN", ("Hold, not increase: an athlete-reported health problem is present; "
                                "keep the load and monitor."), reasons
    return decision, strat, reasons


# ----------------------------------------------------------------------------
# Extended variables derived from the same inputs
#   EWMA ACWR (Williams 2017), weekly load ramp (Gabbett 2016),
#   training stress balance, Hooper total, sleep index, readiness score
# ----------------------------------------------------------------------------
def ewma_acwr(series, ref, lookback=42, acute_span=7, chronic_span=28):
    days = [ref - timedelta(days=i) for i in range(lookback - 1, -1, -1)]
    loads = [series[d]["load"] if d in series else 0.0 for d in days]
    a, cch = 2.0 / (acute_span + 1), 2.0 / (chronic_span + 1)
    ea = ec = None
    for L in loads:
        ea = L if ea is None else a * L + (1 - a) * ea
        ec = L if ec is None else cch * L + (1 - cch) * ec
    return round(ea / ec, 2) if ec and ec > 0 else None


def weekly_ramp(series, ref):
    this = sum(window_loads(series, ref, 7))
    last = sum(window_loads(series, ref - timedelta(days=7), 7))
    return round((this - last) / last * 100.0, 1) if last > 0 else None


def training_stress_balance(series, ref):
    acute = window_loads(series, ref, 7); chronic = window_loads(series, ref, 28)
    am = sum(acute) / len(acute) if acute else 0.0
    cm = sum(chronic) / len(chronic) if chronic else 0.0
    return round(cm - am, 1)   # positive = freshening, negative = fatiguing


def hooper_total(comp, ref):
    t = comp.get(ref, {})
    parts = []
    if t.get("fatigue") is not None: parts.append(t["fatigue"])
    if t.get("stress") is not None: parts.append(t["stress"])
    if t.get("soreness") is not None: parts.append(t["soreness"])
    if t.get("sleep") is not None: parts.append(t["sleep"])       # sleep already 1=best..7=worst
    return round(sum(parts), 1) if parts else None                # range 4..28, higher = worse


def sleep_index(comp, ref):
    today = comp.get(ref, {}).get("sleep")
    week = _recent(comp, ref, 7, "sleep")
    mean = round(sum(week) / len(week), 1) if week else None
    debt = bool(mean is not None and mean >= 4.5)   # sleep 1=very good..7=very poor
    return {"today": today, "week_mean": mean, "debt": debt}


def readiness_score(m):
    """0-100 readiness, blending wellness vs baseline with the workload ratio. Documented heuristic."""
    s = 60.0
    if m.get("wellness_z") is not None:
        s += 12.0 * max(-2.0, min(2.0, m["wellness_z"]))
    if m.get("acwr") is not None:
        if m["acwr"] > 1.3: s -= 18.0 * (m["acwr"] - 1.3)
        if m["acwr"] < 0.8: s -= 10.0 * (0.8 - m["acwr"])
    if m.get("monotony") is not None and m["monotony"] >= 2.0:
        s -= 8.0
    return round(max(0.0, min(100.0, s)), 0)


def extended_metrics(series, comp, ref, m):
    si = sleep_index(comp, ref)
    return {
        "ewma_acwr": ewma_acwr(series, ref),
        "weekly_ramp_pct": weekly_ramp(series, ref),
        "tsb": training_stress_balance(series, ref),
        "hooper_total": hooper_total(comp, ref),
        "sleep_today": si["today"], "sleep_week_mean": si["week_mean"], "sleep_debt": si["debt"],
        "readiness": readiness_score(m),
    }


# ----------------------------------------------------------------------------
# Reporting cadence: daily 20:00, weekly (Saturday), monthly, quarterly
# ----------------------------------------------------------------------------
def reports_due(ref):
    """Which reports the agent issues on this date. Daily always; weekly on Saturday;
    monthly on the last Saturday of the month; quarterly on the last Saturday of Mar/Jun/Sep/Dec."""
    due = ["daily"]
    if ref.weekday() == 5:  # Saturday
        due.append("weekly")
        if (ref + timedelta(days=7)).month != ref.month:  # last Saturday of the month
            due.append("monthly")
            if ref.month in (3, 6, 9, 12):
                due.append("quarterly")
    return due


def _period_table(recs, athletes, ref, days):
    rows = []
    for a in athletes:
        series = daily_series(recs, a)
        comp = component_daily(recs, a)
        win = [series[d]["load"] for d in series if ref - timedelta(days=days - 1) <= d <= ref]
        total = round(sum(win), 0)
        sl = _recent(comp, ref, days, "sleep"); fa = _recent(comp, ref, days, "fatigue")
        half = days // 2
        first = [series[d]["load"] for d in series if ref - timedelta(days=days - 1) <= d <= ref - timedelta(days=half)]
        second = [series[d]["load"] for d in series if ref - timedelta(days=half - 1) <= d <= ref]
        trend = "up" if sum(second) > sum(first) * 1.1 else "down" if sum(second) < sum(first) * 0.9 else "stable"
        rows.append((a, total, round(sum(sl) / len(sl), 1) if sl else "-", round(sum(fa) / len(fa), 1) if fa else "-", trend))
    return rows


def build_periodic(recs, athletes, ref, due):
    html_parts, text_parts = [], []
    for tag, days, title in (("weekly", 7, "Weekly review (last 7 days)"),
                             ("monthly", 28, "Monthly review (last 28 days)")):
        if tag in due:
            rows = _period_table(recs, athletes, ref, days)
            t = [f"\n=== {title} ==="] + [f"{a}: load {tot} AU | sleep {sl}/7 | fatigue {fa}/7 | trend {tr}"
                                          for a, tot, sl, fa, tr in rows]
            text_parts.append("\n".join(t))
            tr_rows = "".join(f"<tr><td>{a}</td><td>{tot}</td><td>{sl}</td><td>{fa}</td><td>{tr}</td></tr>"
                              for a, tot, sl, fa, tr in rows)
            html_parts.append(f'<h3 style="color:#1F3864">{title}</h3><table cellpadding="5" style="border-collapse:collapse;font-size:13px">'
                              f'<tr style="background:#DDEBF7"><th>Athlete</th><th>Load (AU)</th><th>Sleep</th><th>Fatigue</th><th>Trend</th></tr>{tr_rows}</table>')
    if "quarterly" in due:
        rows = _period_table(recs, athletes, ref, 90)
        lines = ["\n=== Quarterly review and next-mesocycle draft (last 90 days) ==="]
        hrows = ""
        for a, tot, sl, fa, tr in rows:
            if tr == "up":
                meso = "consider a recovery or deload block next"
            elif tr == "down" and (fa == "-" or (isinstance(fa, float) and fa <= 4)):
                meso = "room for a build block next"
            else:
                meso = "hold the current block"
            lines.append(f"{a}: 90-day load {tot} AU | trend {tr} | {meso}")
            hrows += f"<tr><td>{a}</td><td>{tot}</td><td>{tr}</td><td>{meso}</td></tr>"
        text_parts.append("\n".join(lines))
        html_parts.append('<h3 style="color:#1F3864">Quarterly review and next-mesocycle draft (last 90 days)</h3>'
                          '<table cellpadding="5" style="border-collapse:collapse;font-size:13px">'
                          '<tr style="background:#DDEBF7"><th>Athlete</th><th>90-day load</th><th>Trend</th><th>Mesocycle suggestion</th></tr>'
                          f'{hrows}</table>')
    return ("".join(html_parts), "\n".join(text_parts))


# ----------------------------------------------------------------------------
# Orchestration
# ----------------------------------------------------------------------------
def run(input_csv, ref_date=None, out_html=None, audit_log=None, mode="auto"):
    recs = load_records(input_csv)
    dates = [r["date"] for r in recs if r["date"]]
    if ref_date is None:
        ref_date = max(dates) if dates else date.today()
    athletes = sorted({r["athlete"] for r in recs if r["athlete"]})
    results = []
    for a in athletes:
        series = daily_series(recs, a)
        if ref_date not in series:
            continue  # no entry for the reference day
        m = compute_metrics(series, ref_date)
        if m is None:
            continue
        decision, reasons = decide(m)
        comp = component_daily(recs, a)
        reasons = reasons + wellness_flags(comp, ref_date)
        sr = soreness_reason(comp, series, ref_date)
        if sr:
            reasons.append(sr)
        ext = extended_metrics(series, comp, ref_date, m)
        strat = strategy(decision, m)
        hs = health_status(recs, a, ref_date)
        decision, strat, reasons = apply_health_override(decision, strat, reasons, hs)
        advice = wellness_advice(reasons)
        results.append({"athlete": a, "metrics": m, "decision": decision, "reasons": reasons,
                        "strategy": strat, "advice": advice, "health": hs, "ext": ext})
    html, text = build_report(results, ref_date)
    due = reports_due(ref_date) if mode == "auto" else (["daily"] if mode == "daily" else ["daily", mode])
    extra_html, extra_text = build_periodic(recs, athletes, ref_date, due)
    html = html.replace("</body>", extra_html + "</body>") if "</body>" in html else html + extra_html
    text = text + ("\n" + extra_text if extra_text else "")
    text += "\n\nReports issued: " + ", ".join(due)
    if out_html:
        with open(out_html, "w", encoding="utf-8") as f:
            f.write(html)
    if audit_log:
        append_audit(audit_log, results, ref_date)
    return results, html, text


def main():
    ap = argparse.ArgumentParser(description="HBD training-load monitoring agent")
    ap.add_argument("--input", required=True, help="responses CSV exported from the form sheet")
    ap.add_argument("--date", help="reference day YYYY-MM-DD (default: latest in data)")
    ap.add_argument("--out", default="hbd_report.html", help="HTML report path")
    ap.add_argument("--log", default="hbd_audit.csv", help="audit log CSV path")
    ap.add_argument("--mode", default="auto", choices=["auto", "daily", "weekly", "monthly", "quarterly"],
                    help="auto issues daily plus any due weekly/monthly/quarterly report")
    args = ap.parse_args()
    ref = _parse_date(args.date) if args.date else None
    results, html, text = run(args.input, ref, args.out, args.log, args.mode)
    print(text)
    print(f"\nHTML report: {args.out}\nAudit log: {args.log}")


if __name__ == "__main__":
    main()
